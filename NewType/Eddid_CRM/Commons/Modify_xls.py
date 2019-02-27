#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2019-02-19 17:41:50
# @Author  : Onedi (Onedi@qq.com)
# @Link    : ${link}
# @Version : $Id$

from openpyxl import load_workbook
from xlutils.copy import copy
import os
import random

class Modifyxls():

    def __init__(self, file_url):
        self.path = file_url
        # self.workbook = xlrd.open_workbook(self.path)
        self.workbook = load_workbook(self.path)

    def readxls(self):

        # 打开excel文件,open_workbook(path),path为excel所在的路径
        # 打开excel表,这里表示打开第一张表
        # table = self.workbook.sheets()[0]
        table = self.workbook.active
        # nrows = table.nrows     # 获取excel的行数
        nrows = table.max_row
        # print(nrows)
        # ncols = table.ncols     #获取excel的列数
        ncols = table.max_column
        # print(ncols)
        # keys = table.row_values(0)      #获取第一行的值

        keys = [keys for keys in list(table.rows)[0]]
        # print(keys[0].value)

        Data = []       #创建一个list，用于存放
        x = 1
        for i in range(nrows-1):
            s = {}
            # print(i)
            # values = table.row_values(x)
            values = [keys for keys in list(table.rows)[x]]
            # print(values)
            for j in range(ncols):
                # print('j=',j)
                s[keys[j].value] = values[j].value
            # print(s)
            Data.append(s)
            x += 1

        return Data

    def writexls(self):

        sheet = self.workbook.active
        id_code = random.randint(100000,135483216542154)
        email = 'onedi2%s@qq.com' %(random.randint(0,4541545))
        sheet.cell(row=2, column=13).value = id_code
        sheet.cell(row=2, column=15).value = email
        self.workbook.save(self.path)
        return id_code

        # wb = copy(self.workbook)
        # sheet = wb.active()
        # id_code = random.randint(100000,135483216542154)
        # email = 'onedi2%s' %(random.randint(4541545))
        # sheet.write(1, 12, id_code)
        # sheet.write(1, 14, email)
        # wb.save(path)
        # return id_code



if __name__ == "__main__":

    file_url = os.path.abspath(os.path.dirname(os.getcwd()))+'/config/Ayers1.xlsx'

    modify = Modifyxls(file_url)
    data = modify.readxls()
    print(data)
    for res in data:
        print(int(res['id_code']))
    #     print(res)
    # code = modify.writexls(file_url)
    # print(code)